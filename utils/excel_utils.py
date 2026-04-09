import io
from openpyxl import Workbook, load_workbook

def export_to_excel(columns, data, sheet_name='Sheet1'):
    """Export data to Excel and return bytes"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col_idx, header in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def import_from_excel(file):
    """Import data from Excel file and return list of dictionaries"""
    wb = load_workbook(file, data_only=True)  # data_only=True to get calculated values instead of formulas
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else '')

    # Header mapping: normalize Chinese headers to English keys
    header_map = {
        '编码': 'code',
        '物料名称': 'name', '货品名称': 'name', '物品名称': 'name', '名称': 'name',
        '规格型号': 'spec', '规格': 'spec',
        '单位': 'unit',
        '分类代码': 'category_code', '分类': 'category_code',
        '厂商': 'manufacturer',
        '保存条件': 'storage_condition',
        '保质期': 'shelf_life', '保质期天': 'shelf_life',
        '备注': 'remark',
        '物料编码': 'material_code', '编码': 'material_code',
        '数量': 'quantity', '库存': 'quantity',
        '批次号': 'batch_no', '批次': 'batch_no',
        '生产日期': 'production_date',
        '有效日期': 'expiry_date', '有效期': 'expiry_date'
    }

    # Read data rows
    data = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx in range(len(headers)):
            header = headers[col_idx]
            if not header:
                continue
            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
            # Map header to normalized key
            normalized_key = header_map.get(header, header)
            if normalized_key:
                row_data[normalized_key] = cell_value
        # Skip empty rows
        if any(v is not None for v in row_data.values()):
            data.append(row_data)

    return data

def import_from_excel_by_position(file, num_columns, skip_header=True):
    """Import data from Excel file using position-based columns
    Returns list of dicts with keys 'col1', 'col2', etc."""
    wb = load_workbook(file)
    ws = wb.active

    # Start from row 2 if skip_header, otherwise from row 1
    start_row = 2 if skip_header else 1
    data = []
    for row_idx in range(start_row, ws.max_row + 1):
        row_data = {}
        has_value = False
        for col_idx in range(1, num_columns + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[f'col{col_idx}'] = cell_value
            if cell_value is not None:
                has_value = True
        # Skip empty rows
        if has_value:
            data.append(row_data)

    return data
